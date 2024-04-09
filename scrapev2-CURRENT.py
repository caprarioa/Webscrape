import random
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait


def get_random_user_agent():
    user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36"
    ]
    return random.choice(user_agents)

def configure_firefox_browser(user_agent):
    options = Options()
    options.headless = False
    options.add_argument("user-agent=" + user_agent)
    return webdriver.Firefox(options=options)

def search_part(browser, part):
    search_box = WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.XPATH, "//input[@id='main_search_5']")))
    search_box.send_keys(Keys.CONTROL + "a", Keys.BACKSPACE)
    time.sleep(1)
    search_box.send_keys(part, Keys.RETURN)

def get_description(browser):
    try:
        WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.XPATH, "//div[@class='page-builder-layout-module    desktop-only']//div[@class='product-title-module']//h1[@class='product-title']")))
        description = browser.find_element(By.XPATH, "//div[@class='page-builder-layout-module    desktop-only']//div[@class='product-title-module']//h1[@class='product-title']")
        return description.text
    except:
        return "Desc not found"
    
def get_yearmodel(browser):
    try:
        WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.XPATH, "//div[@class='page-builder-layout-module    desktop-only']//div[@class='product-title-module']//p[@class='product-subtitle']")))
        yearmodel = browser.find_element(By.XPATH, "//div[@class='page-builder-layout-module    desktop-only']//div[@class='product-title-module']//p[@class='product-subtitle']")
        return yearmodel.text
    except:
        return "Year and model not found"

def main():
    user_agent = get_random_user_agent()
    browser = configure_firefox_browser(user_agent)
    browser.get("https://www.moparsalesdirect.com/oem-parts/mopar-cam-phaser-actuator-5184101ah")

    wb = openpyxl.load_workbook('parts_search.xlsx')
    sheet = wb['search']

    for i in range(1, sheet.max_row + 1):
        part = sheet.cell(row=i, column=1).value
        if part:
            search_part(browser, part)
            time.sleep(1)
            description = get_description(browser)
            sheet.cell(row=i, column=3).value = description
            print (description)
            time.sleep(1)
            yearmodel = get_yearmodel(browser)
            sheet.cell(row=i, column=4).value = yearmodel
            time.sleep(1)
            wb.save('parts_search.xlsx')
            time.sleep(1)

    wb.close()
    browser.close()

if __name__ == "__main__":
    main()
